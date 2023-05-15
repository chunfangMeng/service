import os
import xlrd

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class ExcelHandler(object):
    def __init__(self, file_path: str):
        self.file_path = file_path

    @classmethod
    def check_sheet(cls, wb, sheet_name):
        """
        简单检查工作簿是否存在
        :param wb:
        :param sheet_name:
        :return:
        """
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError()
            return True, None
        except ValueError:
            return False, 'sheet_name不存在'

    def _xls_format_read(self, sheet_name=None, is_read_header=False):
        """
        xls文件读取
        :param sheet_name: 活动工作簿
        :param min_row: 从第几行开始
        :return:
        """
        wb = xlrd.open_workbook(self.file_path)
        if sheet_name is not None:
            sheet = wb.sheet_by_name(sheet_name)
        else:
            sheet = wb.sheet_by_index(0)
        rows = []
        for i in range(0 if is_read_header else 1, sheet.nrows):
            rows.append([col.value for col in sheet.row(i)])
        return rows, None

    def _xlsx_format_read(self, sheet_name=None, is_read_header=False):
        """
        xlsx文件读取
        :param sheet_name: 活动工作簿
        :param is_read_header:
        :return:
        """
        wb = load_workbook(filename=self.file_path)
        if sheet_name is None:
            sheet = wb.active
        else:
            is_valid, msg = self.check_sheet(wb, sheet_name)
            if not is_valid:
                return [], msg
            sheet = wb[sheet_name]
        if not is_read_header:
            min_row = 2
        rows = []
        for row in sheet.iter_rows(min_row=min_row):
            row_data = list()
            for cell in row:
                if cell.value is not None:
                    row_data.append(cell.value)
            if len(row_data) == 0:
                continue
            rows.append(row_data)
        return rows, None

    def read_excel(self, sheet_name=None, is_read_header=False):
        """
        读取excel
        :param sheet_name: 工作簿名，不传sheet_name时默认是读取活动工作簿
        :param is_read_header:
        :return:
        """
        if os.path.splitext(self.file_path)[1] == '.xls':
            rows, _ = self._xls_format_read(sheet_name, is_read_header)
        elif os.path.splitext(self.file_path)[1] == '.xlsx':
            rows, _ = self._xlsx_format_read(sheet_name, is_read_header)
        return rows, None

    def write_excel(self, excel_data, is_overwrite=False):
        """
        写excel
        :param excel_data:
        :param is_overwrite: 是否覆盖源文件
        :return:
        """
        wb = Workbook()
        sheet = wb.active
        for row_index, row_data in enumerate(excel_data):
            for column_index, value in enumerate(row_data):
                column_letter = get_column_letter(column_index + 1)
                sheet[f'{column_letter}{row_index + 1}'] = value
        self._save_write_file(wb, is_overwrite)

    def _save_write_file(self, wb: Workbook, is_overwrite=False):
        """
        写文件时保存文件
        :param wb: Workbook
        :param is_overwrite:
        :return:
        """
        if is_overwrite:
            wb.save(filename=self.file_path)
        if os.path.isfile(self.file_path):
            full_file_name = os.path.basename(self.file_path)
            full_name_list = os.path.splitext(full_file_name)
            start_index = full_name_list[0].find('(')
            end_index = full_name_list[0].rfind(')')
            if start_index != -1 and end_index != -1:
                count = int(full_name_list[0][start_index+1:end_index])
                file_name = full_name_list[0][:start_index] + '(' + str(count) + ')'
            else:
                file_name = f'{full_name_list[0]}_(1)'
            new_file_name = f'{file_name}.{full_name_list[-1]}'
            wb.save(filename=new_file_name)

